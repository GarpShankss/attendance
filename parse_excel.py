"""
Step 1: Parse messy multi-sheet excel files (xls/xlsx) into clean
list-of-dict rows per sheet, auto-detecting the header (handles the
2-row 'group header + column name' pattern seen in these payroll sheets).

Usage:
    python3 parse_excel.py <path1> <path2> ...
Outputs one JSON file per input next to this script: <name>_parsed.json
{ "sheet_name": { "columns": [...], "rows": [ {col: val, ...}, ... ] } }
"""
import sys, json, re
import openpyxl
import xlrd

SLNO_PATTERN = re.compile(r"^\s*s\.?\s*l?\.?\s*no\.?\s*$", re.I)
EMP_ID_PATTERN = re.compile(r"^\s*(emp\.?\s*id|id\s*no\.?|employee\s*id|emp\s*name|employee\s*name)\s*$", re.I)
TOTAL_PATTERN = re.compile(r"total", re.I)


def is_blank(v):
    return v is None or (isinstance(v, str) and v.strip() == "")


def get_rows_xlsx(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    out = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = [[_clean_cell(v) for v in r]
                for r in ws.iter_rows(values_only=True)]
        out[sn] = rows
    return out


def _clean_cell(val):
    """
    Normalise a cell value coming from Excel:
    - Floats that are very close to an integer (within 1e-6) are rounded to int.
      This handles formula cells like =13354*30/30 that openpyxl reads as
      13353.999999 or 13354.000001 instead of the displayed integer 13354.
    - Genuine decimals (e.g. 28.5 pay days) are left as-is.
    """
    if isinstance(val, float):
        rounded = round(val)
        if abs(val - rounded) < 1e-6:
            return rounded        # treat as integer
    return val


def get_rows_xls(path):
    wb = xlrd.open_workbook(path)
    out = {}
    for sn in wb.sheet_names():
        sh = wb.sheet_by_name(sn)
        rows = []
        for r in range(sh.nrows):
            row = []
            for c in range(sh.ncols):
                cell = sh.cell(r, c)
                if cell.ctype == xlrd.XL_CELL_DATE:
                    # Only convert to datetime if value looks like a real calendar
                    # date (serial > 367 = after 1901-01-01).
                    # Small integers like 30 / 31 (working days) are plain numbers
                    # that happen to share a cell type due to cell formatting.
                    if cell.value > 367:
                        try:
                            val = xlrd.xldate_as_datetime(cell.value, wb.datemode)
                        except Exception:
                            val = _clean_cell(cell.value)
                    else:
                        val = _clean_cell(cell.value)
                else:
                    val = _clean_cell(cell.value)
                row.append(val)
            rows.append(row)
        out[sn] = rows
    return out


def find_header_row(rows):
    """Find the row whose first few cells look like 'Sl.No' / 'S.No' / 'Emp ID' style labels."""
    for i, row in enumerate(rows[:25]):
        for cell in row[:5]:
            if isinstance(cell, str):
                cleaned = cell.strip()
                if SLNO_PATTERN.match(cleaned) or EMP_ID_PATTERN.match(cleaned):
                    return i
    return None


import datetime as dt

DATE_LIKE = re.compile(
    r"^\d{4}-\d{2}-\d{2}"   # 2026-05-01 ...
    r"|^\d{1,2}/\d{1,2}/\d{2,4}"  # 1/5/2026
    r"|^\d{1,2}-[A-Za-z]{3}-\d{2,4}"  # 1-May-2026
)


def _is_date_like(v):
    if isinstance(v, (dt.datetime, dt.date)):
        return True
    if isinstance(v, str) and DATE_LIKE.match(v.strip()):
        return True
    return False


def ffill(row):
    """Forward-fill non-blank values, but skip date-like values."""
    out, last = [], None
    for v in row:
        if not is_blank(v) and not _is_date_like(v):
            last = v
        out.append(last)
    return out


def build_columns(rows, header_idx):
    below = rows[header_idx + 1] if header_idx + 1 < len(rows) else None
    case_a = below is not None and is_blank(below[0]) and any(not is_blank(v) for v in below)

    if case_a:
        top = ffill(rows[header_idx])
        sub = below
        data_start = header_idx + 2
    else:
        top = ffill(rows[header_idx - 1]) if header_idx > 0 else [None] * len(rows[header_idx])
        sub = rows[header_idx]
        data_start = header_idx + 1

    cols = []
    for i in range(len(sub)):
        t = str(top[i]).strip() if not is_blank(top[i]) and not _is_date_like(top[i]) else None
        s = str(sub[i]).strip() if not is_blank(sub[i]) else None
        if t and s and t != s:
            cols.append(f"{t} - {s}")
        elif s:
            cols.append(s)
        elif t:
            cols.append(t)
        else:
            cols.append(f"col_{i + 1}")

    # de-duplicate: use _2, _3 ... suffix instead of (2), (3)
    # Also sanitize: collapse any whitespace (newlines, tabs) to single space
    seen = {}
    final = []
    for c in cols:
        c = re.sub(r'[\r\n\t]+', ' ', c).strip()
        c = re.sub(r' {2,}', ' ', c)
        key = c.lower()
        seen[key] = seen.get(key, 0) + 1
        col_name = c if seen[key] == 1 else f"{c}_{seen[key]}"
        if col_name.lower() in ("net pay_2", "net pay(2)", "net_pay(2)"):
            col_name = "Phone Number"
        final.append(col_name)
    return final, data_start


def extract_data_rows(rows, data_start, columns):
    data = []
    for row in rows[data_start:]:
        if all(is_blank(v) for v in row):
            break
        first_text = next((str(v) for v in row[:2] if not is_blank(v)), "")
        if TOTAL_PATTERN.search(first_text) and not isinstance(row[0], (int, float)):
            break
        if is_blank(row[0]):
            continue  # stray subtotal/blank line with no serial number
        record = {}
        for i, col in enumerate(columns):
            val = row[i] if i < len(row) else None
            record[col] = val if not is_blank(val) else None
        if any(v is not None for v in record.values()):
            data.append(record)
    return data


def parse_workbook(path):
    rows_by_sheet = get_rows_xlsx(path) if path.lower().endswith("xlsx") else get_rows_xls(path)
    result = {}
    for sn, rows in rows_by_sheet.items():
        header_idx = find_header_row(rows)
        if header_idx is None:
            result[sn] = {"columns": [], "rows": [], "note": "header not detected"}
            continue
        columns = build_columns(rows, header_idx)
        data = extract_data_rows(rows, columns[1], columns[0])
        result[sn] = {"columns": columns[0], "rows": data}
    return result


def drop_empty_columns(columns, data):
    keep = [c for c in columns if any(row.get(c) is not None for row in data)]
    cleaned = [{c: row[c] for c in keep} for row in data]
    return keep, cleaned


if __name__ == "__main__":
    for path in sys.argv[1:]:
        parsed = parse_workbook(path)
        for sn, d in parsed.items():
            d["columns"], d["rows"] = drop_empty_columns(d["columns"], d["rows"])
        out_name = path.rsplit("/", 1)[-1].rsplit(".", 1)[0] + "_parsed.json"
        with open(out_name, "w") as f:
            json.dump(parsed, f, indent=2, default=str)
        print(f"\n=== {path} -> {out_name} ===")
        for sn, d in parsed.items():
            print(f"  [{sn}] {len(d['columns'])} cols, {len(d['rows'])} rows")
