"""
Salary Fitment Letter Generator and Dispatcher.
Generates Excel (.xlsx) and PDF (.pdf) matching the format in 'Fitment letter.xlsx',
and dispatches to employees via WhatsApp.
"""
import os
import io
import datetime
import tempfile

if not os.environ.get("PLAYWRIGHT_BROWSERS_PATH"):
    win_path = r"C:\Users\Sellogs\AppData\Local\ms-playwright"
    if os.path.exists(win_path):
        os.environ["PLAYWRIGHT_BROWSERS_PATH"] = win_path

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from jinja2 import Environment, FileSystemLoader
from playwright.sync_api import sync_playwright
from whatsapp_utils import send_payslip_whatsapp

TEMPLATE_XLSX_PATH = os.path.join(os.path.dirname(__file__), "templates", "fitment_letter_template.xlsx")


def calculate_fitment_side(basic: float, da: float, hra: float, location: str = "Bangalore", warehouse: str = "") -> dict:
    """
    Calculate full breakdown for one side (Previous or New) of fitment comparison.
    """
    basic = float(basic or 0)
    da = float(da or 0)
    hra = float(hra or 0)
    
    # Statutory Bonus @ 8.33% of (Basic + DA)
    bonus = round((basic + da) * 0.0833, 2)
    gross = round(basic + da + hra + bonus, 2)
    
    # PF Base
    is_abb = bool(warehouse and ("abb" in str(warehouse).lower()))
    if is_abb:
        pf_base = min(basic + da, 15000.0)
    else:
        pf_base = basic + da
        
    pf_employee = round(pf_base * 0.12, 2)
    pf_employer = round(pf_base * 0.13, 2)
    
    # ESI (ceiling 21,000 / 20,999)
    if gross < 21000:
        esic_employee = round(gross * 0.0075, 2)
        esic_employer = round(gross * 0.0325, 2)
    else:
        esic_employee = 0.0
        esic_employer = 0.0
        
    # PT
    loc_lower = str(location or "").lower()
    if "hyderabad" in loc_lower:
        if gross >= 20000:
            pt = 200.0
        elif gross >= 15000:
            pt = 150.0
        else:
            pt = 0.0
    else:
        pt = 200.0 if gross >= 25000 else 0.0
        
    total_deduction = round(pf_employee + esic_employee + pt, 2)
    net_take_home = round(gross - total_deduction, 2)
    monthly_take_home = net_take_home
    final_ctc = round(gross + pf_employer + esic_employer, 2)
    
    return {
        "basic": basic,
        "da": da,
        "hra": hra,
        "bonus": bonus,
        "gross": gross,
        "pf_employee": pf_employee,
        "esic_employee": esic_employee,
        "pt": pt,
        "total_deduction": total_deduction,
        "net_take_home": net_take_home,
        "monthly_take_home": monthly_take_home,
        "pf_employer": pf_employer,
        "esic_employer": esic_employer,
        "final_ctc": final_ctc
    }


def generate_fitment_excel(data: dict) -> bytes:
    """
    Populates template xlsx and returns byte content.
    """
    if os.path.exists(TEMPLATE_XLSX_PATH):
        wb = openpyxl.load_workbook(TEMPLATE_XLSX_PATH)
    else:
        # Fallback to root if template not yet copied
        root_path = os.path.join(os.path.dirname(__file__), "Fitment letter.xlsx")
        wb = openpyxl.load_workbook(root_path)

    ws = wb.active

    company_name = data.get("company_name", "RS MAN- TECH").strip()
    emp_name = data.get("emp_name", "").strip()
    designation = data.get("designation", "").strip()
    department = data.get("department", data.get("warehouse", "")).strip()
    emp_id = data.get("emp_id", "").strip()
    doj = data.get("doj", "").strip()
    location = data.get("location", "Bangalore").strip()
    date_str = data.get("date_str", datetime.date.today().strftime("%d.%m.%Y"))
    
    prev_year_label = data.get("prev_year_label", "25-26").strip()
    new_year_label = data.get("new_year_label", "26-27").strip()
    
    prev_basic = float(data.get("prev_basic", 0))
    prev_da = float(data.get("prev_da", 0))
    prev_hra = float(data.get("prev_hra", 0))
    
    new_basic = float(data.get("new_basic", 0))
    new_da = float(data.get("new_da", 0))
    new_hra = float(data.get("new_hra", 0))

    # Header and Company
    ws["B5"] = f" {company_name} "
    ws["B7"] = "SALARY FITMENT LETTER"
    
    # Employee info
    ws["B10"] = "Name:                                               "
    ws["C10"] = emp_name
    ws["B11"] = "Designation:                                 "
    ws["C11"] = designation
    ws["B12"] = "Department:                                  "
    ws["C12"] = department
    ws["B15"] = "Emp Code No:                                   "
    ws["C15"] = emp_id
    ws["B16"] = "Date of Joining:                                 "
    ws["C16"] = doj
    
    # Table column labels
    ws["B18"] = "Particulars"
    ws["C18"] = prev_year_label
    ws["D18"] = new_year_label
    
    # Inputs
    ws["C19"] = prev_basic
    ws["D19"] = new_basic
    
    ws["C20"] = prev_da
    ws["D20"] = new_da
    
    ws["C21"] = prev_hra
    ws["D21"] = new_hra
    
    # Formulas are already in C22:D31 of the template!
    # Ensure standard formulas are maintained
    ws["C22"] = "=(C19+C20)*8.33%"
    ws["D22"] = "=(D19+D20)*8.33%"
    
    ws["C23"] = "=SUM(C19:C22)"
    ws["D23"] = "=SUM(D19:D22)"
    
    ws["C24"] = "=(C19+C20)*12%"
    ws["D24"] = "=(D19+D20)*12%"
    
    ws["C25"] = "=+IF(C23<20999,C23*0.75%,0)"
    ws["D25"] = "=+IF(D23<20999,D23*0.75%,0)"
    
    ws["C26"] = "=ROUND(IF(C23>=25000,200,0),0)"
    ws["D26"] = "=ROUND(IF(D23>=25000,200,0),0)"
    
    ws["C27"] = "=C23-C24-C25-C26"
    ws["D27"] = "=D23-D24-D25-D26"
    
    ws["C28"] = "=C27"
    ws["D28"] = "=D27"
    
    ws["C29"] = "=(C19+C20)*13%"
    ws["D29"] = "=(D19+D20)*13%"
    
    ws["C30"] = "=+IF(C23<20999,C23*3.25%,0)"
    ws["D30"] = "=+IF(D23<20999,D23*3.25%,0)"
    
    ws["C31"] = "=C23+C29+C30"
    ws["D31"] = "=D23+D29+D30"

    # Consent and footer
    ws["B33"] = f"I {emp_name or '______________'} agree to the above salary fitment with my full consent"
    ws["B34"] = " and accept the Increment Letter."
    ws["B35"] = "Bonus will be add in your monthly salary."
    ws["B38"] = f"Place: {location or 'Bangalore'}"
    ws["B39"] = f"Date - {date_str}"
    ws["D39"] = "Signature"

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream.getvalue()


def _launch_browser(p):
    try:
        return p.chromium.launch(headless=True)
    except Exception:
        try:
            return p.chromium.launch(headless=True, channel="chrome")
        except Exception:
            return p.chromium.launch(headless=True, channel="msedge")


def generate_fitment_pdf(data: dict) -> bytes:
    """
    Renders the Fitment Letter HTML template and converts to PDF using Playwright.
    """
    location = data.get("location", "Bangalore")
    warehouse = data.get("warehouse", data.get("department", ""))
    
    prev_calc = calculate_fitment_side(
        basic=data.get("prev_basic", 0),
        da=data.get("prev_da", 0),
        hra=data.get("prev_hra", 0),
        location=location,
        warehouse=warehouse
    )
    new_calc = calculate_fitment_side(
        basic=data.get("new_basic", 0),
        da=data.get("new_da", 0),
        hra=data.get("new_hra", 0),
        location=location,
        warehouse=warehouse
    )

    template_dir = os.path.join(os.path.dirname(__file__), "templates")
    env = Environment(loader=FileSystemLoader(template_dir))
    template = env.get_template("fitment_letter_template.html")
    
    import base64
    logo_path = os.path.join(os.path.dirname(__file__), "logo.png")
    logo_b64 = ""
    if os.path.exists(logo_path):
        with open(logo_path, "rb") as f:
            logo_b64 = base64.b64encode(f.read()).decode("utf-8")

    html_out = template.render(
        company_name=data.get("company_name", "RS MAN- TECH"),
        emp_name=data.get("emp_name", ""),
        designation=data.get("designation", ""),
        department=data.get("department", data.get("warehouse", "")),
        emp_id=data.get("emp_id", ""),
        doj=data.get("doj", ""),
        location=location,
        date_str=data.get("date_str", datetime.date.today().strftime("%d.%m.%Y")),
        prev_year_label=data.get("prev_year_label", "25-26"),
        new_year_label=data.get("new_year_label", "26-27"),
        prev=prev_calc,
        new=new_calc,
        logo_base64=logo_b64
    )

    with sync_playwright() as p:
        browser = _launch_browser(p)
        page = browser.new_page()
        page.set_content(html_out, wait_until="networkidle")
        pdf_bytes = page.pdf(
            format="A4",
            print_background=True,
            margin={"top": "0.4in", "bottom": "0.4in", "left": "0.5in", "right": "0.5in"}
        )
        browser.close()

    return pdf_bytes


def send_fitment_letter_whatsapp(phone_number: str, data: dict) -> bool:
    """
    Generates Fitment Letter PDF and sends it via WhatsApp.
    """
    emp_name = data.get("emp_name", "Employee")
    pdf_bytes = generate_fitment_pdf(data)
    pdf_filename = f"Fitment_Letter_{emp_name.replace(' ', '_')}_{data.get('new_year_label', '26-27')}.pdf"
    
    month_desc = f"Fitment Letter ({data.get('prev_year_label', '25-26')} to {data.get('new_year_label', '26-27')})"
    return send_payslip_whatsapp(
        phone_number=phone_number,
        emp_name=emp_name,
        month=month_desc,
        pdf_bytes=pdf_bytes,
        pdf_filename=pdf_filename
    )
