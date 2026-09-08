"""
FastAPI backend.

  GET    /warehouses                   {location: [warehouse, ...]} - the filter/assignment options
  POST   /upload/preview               upload a file -> parses it, returns sheet names/row counts + a token
  POST   /upload/confirm               {token, location, mapping:{sheet: warehouse|""}} -> loads into Mongo
  GET    /collections                  list collections, each with its location/warehouse/sheet tag
  GET    /collections/{name}           {columns, rows} - columns read live from the data, nothing hardcoded
  PATCH  /collections/{name}/{row_id}  edit one cell: body {"column":, "value":}
  GET    /data?location=&warehouse=    combined rows across every collection matching that tag

Run:
    uvicorn app:app --reload
Then open http://localhost:8000
"""
import os, shutil, tempfile, uuid
from datetime import datetime
from fastapi import FastAPI, UploadFile, File, HTTPException, Body
from fastapi.staticfiles import StaticFiles
from pymongo import MongoClient
import calendar
import json
import asyncio
from fastapi.responses import StreamingResponse
from datetime import timedelta

class EventBroadcaster:
    def __init__(self):
        self.queues = []
    async def listen(self):
        q = asyncio.Queue()
        self.queues.append(q)
        try:
            while True:
                msg = await q.get()
                yield f"data: {msg}\n\n"
        except asyncio.CancelledError:
            self.queues.remove(q)

broadcaster = EventBroadcaster()

from mongo_loader import load_sheet_into_collection, guess_warehouse
from parse_excel import parse_workbook, drop_empty_columns
from warehouses_config import LOCATIONS
from salary_calc import recalculate
from payroll_settings import get_config
from payroll_settings import get_config, save_config
from payroll_records import generate_monthly_payroll, get_payroll
from payslip_pdf import generate_and_upload_payslip
from whatsapp_utils import send_payslip_whatsapp

MONGO_URI = os.environ.get("MONGO_URI", "mongodb://localhost:27017")
DB_NAME = os.environ.get("DB_NAME", "payroll_db")

import certifi
if MONGO_URI.startswith("mongodb+srv://"):
    client = MongoClient(MONGO_URI, tlsCAFile=certifi.where())
else:
    client = MongoClient(MONGO_URI)
db = client[DB_NAME]

app = FastAPI()

# internal/meta fields - never shown as a data column, never editable
HIDDEN_FIELDS = {"_id", "_row_id", "_source_file", "_sheet", "_location", "_warehouse",
                 "_upload_month", "_upload_year", "_collection", "created_at", "updated_at", "status", "identity"}

# token -> temp file path, held between /upload/preview and /upload/confirm
PENDING_UPLOADS = {}


@app.get("/debug/trace/{name}/{row_id}")
def debug_trace(name: str, row_id: int):
    """
    Full data-flow trace for a single row.
    Shows the exact values at every stage: MongoDB raw -> API response -> recalculate inputs.
    """
    import re as _re
    from salary_calc import COLUMN_MAP, READONLY_FIELDS

    TRACE_COLS = [
        "FIXED - Working Days",
        "FIXED - Basic",
        "FIXED - DA",
        "FIXED - Other Allows",
        "FIXED - Leave With wages",
        "FIXED - Bonus @8.33%",
        "FIXED - HRA",
        "FIXED - Total",
        "ATTENDANCE - Pay Days",
        "ATTENDANCE - Present Days",
    ]

    if name not in db.list_collection_names():
        raise HTTPException(404, "Unknown collection")
    doc = db[name].find_one({"_row_id": row_id})
    if not doc:
        raise HTTPException(404, "Row not found")

    def _norm(k):
        return _re.sub(r'[\s]+', ' ', str(k)).strip()

    # Stage 1: raw MongoDB document
    stage1 = {}
    for col in TRACE_COLS:
        # try exact, then normalised
        val = doc.get(col)
        if val is None:
            for k, v in doc.items():
                if _norm(k) == _norm(col):
                    val = v
                    break
        stage1[col] = {"value": val, "type": type(val).__name__}

    # Stage 2: what the API /data endpoint would send (string conversion happens in JSON)
    # Values go through no transformation — same as stage 1 for numeric fields

    # Stage 3: what recalculate() receives after save_row preprocessing
    def _clean_key(k):
        return _re.sub(r'[\s]+', ' ', str(k)).strip()
    simulated_payload = {_clean_key(k): v for k, v in doc.items()
                         if k not in {"_id", "_row_id", "_source_file", "_sheet",
                                      "_location", "_warehouse", "_upload_month", "_upload_year"}}
    stage3 = {}
    for col in TRACE_COLS:
        val = simulated_payload.get(_clean_key(col))
        stage3[col] = {"value": val, "type": type(val).__name__}

    # Stage 4: what salary_calc.get() reads via COLUMN_MAP
    from salary_calc import get as sc_get, _num
    stage4 = {}
    for logical, col in COLUMN_MAP.items():
        if col in TRACE_COLS:
            stage4[col] = {
                "logical_name": logical,
                "is_readonly": logical in READONLY_FIELDS,
                "value_via_get": sc_get(simulated_payload, logical),
            }

    # Check: are any FIXED columns overwritten by recalculate?
    from salary_calc import recalculate
    result = recalculate(dict(simulated_payload))
    result.pop("__calc_log__", None)
    stage5_changes = {}
    for col in TRACE_COLS:
        before = simulated_payload.get(_clean_key(col))
        after  = result.get(_clean_key(col))
        if before != after:
            stage5_changes[col] = {"before": before, "after": after}

    # All stored keys that look like FIXED columns
    all_fixed_keys = {k: v for k, v in doc.items() if "FIXED" in str(k).upper()}

    return {
        "row_id": row_id,
        "collection": name,
        "stage1_mongodb_raw": stage1,
        "stage3_after_key_clean": stage3,
        "stage4_salary_calc_reads": stage4,
        "stage5_fixed_cols_changed_by_recalculate": stage5_changes,
        "all_fixed_keys_in_db": {str(k): str(v) for k, v in all_fixed_keys.items()},
        "note": "If stage5 shows any changes, recalculate() is overwriting FIXED columns"
    }


@app.get("/debug/row/{name}/{row_id}")
def debug_row(name: str, row_id: int):
    """Show raw stored values + what salary_calc reads for the key input fields."""
    from salary_calc import COLUMN_MAP, get as sc_get, _num
    if name not in db.list_collection_names():
        raise HTTPException(404, "Unknown collection")
    doc = db[name].find_one({"_row_id": row_id})
    if not doc:
        raise HTTPException(404, "Row not found")

    # Show every field that salary_calc cares about
    input_fields = [
        "working_days", "pay_days",
        "fixed_basic", "fixed_da", "fixed_other", "fixed_leave", "fixed_bonus",
        "fixed_hra", "fixed_service_charge", "uniform", "advance",
    ]
    inputs = {}
    for field in input_fields:
        col = COLUMN_MAP.get(field)
        raw = doc.get(col) if col else None
        inputs[field] = {
            "column": col,
            "raw_value": str(raw),
            "raw_type": type(raw).__name__,
            "parsed_num": _num(raw),
        }

    # All keys in the doc (excluding internal)
    all_keys = {k: str(v) for k, v in doc.items() if not k.startswith("_") or k == "_row_id"}
    return {"row_id": row_id, "collection": name, "inputs": inputs, "all_fields": all_keys}


@app.post("/admin/sanitize-keys")
def sanitize_keys():
    """One-time: rename any keys containing newlines/tabs in all collections."""
    import re as _re
    def _clean(k):
        return _re.sub(r'[\s]+', ' ', str(k)).strip()
    total_docs = 0
    for name in db.list_collection_names():
        if name.startswith("system."):
            continue
        for doc in db[name].find({}):
            dirty = {k: v for k, v in doc.items()
                     if k != "_id" and _clean(k) != k}
            if not dirty:
                continue
            # unset old keys, set cleaned keys
            unset = {k: "" for k in dirty}
            setval = {_clean(k): v for k, v in dirty.items()}
            db[name].update_one(
                {"_id": doc["_id"]},
                {"$unset": unset, "$set": setval}
            )
            total_docs += 1
    return {"sanitized_docs": total_docs}


@app.get("/admin/payroll-settings")
def get_payroll_settings():
    return get_config()


@app.put("/admin/payroll-settings")
def update_payroll_settings(payload: dict = Body(...)):
    try:
        saved = save_config(payload)
    except ValueError as e:
        raise HTTPException(400, str(e))
    return saved


@app.post("/admin/backfill-months")
def backfill_months():
    """One-time: stamp _upload_month/_upload_year on all existing docs that lack it."""
    now = datetime.utcnow()
    total = 0
    for name in db.list_collection_names():
        if name.startswith("system."):
            continue
        result = db[name].update_many(
            {"_upload_month": {"$exists": False}},
            {"$set": {"_upload_month": now.month, "_upload_year": now.year}}
        )
        total += result.modified_count
    return {"patched": total}


@app.get("/warehouses")
def get_warehouses():
    return LOCATIONS


@app.post("/upload/preview")
async def upload_preview(file: UploadFile = File(...)):
    suffix = os.path.splitext(file.filename)[1] or ".xlsx"
    fd, tmp_path = tempfile.mkstemp(suffix=suffix)
    os.close(fd)
    with open(tmp_path, "wb") as out:
        shutil.copyfileobj(file.file, out)

    parsed = parse_workbook(tmp_path)
    all_warehouses = sorted({wh for whs in LOCATIONS.values() for wh in whs})
    sheets = []
    for sheet_name, d in parsed.items():
        columns, rows = drop_empty_columns(d["columns"], d["rows"])
        if not rows:
            continue
        sheets.append({
            "sheet_name": sheet_name, "rows": len(rows), "columns": len(columns),
            "guess": guess_warehouse(sheet_name, all_warehouses),
        })

    if not sheets:
        os.remove(tmp_path)
        raise HTTPException(400, "No readable sheets found in that file.")

    token = uuid.uuid4().hex
    PENDING_UPLOADS[token] = {"path": tmp_path, "filename": file.filename}
    return {"token": token, "sheets": sheets}


@app.post("/upload/confirm")
def upload_confirm(payload: dict = Body(...)):
    token = payload.get("token")
    location = payload.get("location")
    mapping = payload.get("mapping", {})  # {sheet_name: warehouse_name_or_empty}
    pending = PENDING_UPLOADS.pop(token, None)
    if not pending:
        raise HTTPException(400, "Unknown or expired upload token - re-upload the file.")

    try:
        parsed = parse_workbook(pending["path"])
        summary = []
        for sheet_name, d in parsed.items():
            warehouse = mapping.get(sheet_name) or None
            result = load_sheet_into_collection(
                sheet_name, d, db, location=location,
                warehouse=warehouse, source_file=pending["filename"],
            )
            if result["rows"]:
                summary.append(result)
    finally:
        os.remove(pending["path"])

    return {"loaded": summary}


@app.get("/collections")
def list_collections():
    names = [n for n in db.list_collection_names() if not n.startswith("system.")]
    out = []
    for name in sorted(names):
        sample = db[name].find_one()
        out.append({
            "name": name,
            "sheet": sample.get("_sheet") if sample else None,
            "location": sample.get("_location") if sample else None,
            "warehouse": sample.get("_warehouse") if sample else None,
            "row_count": db[name].count_documents({}),
        })
    return {"collections": out}


def _flatten_nested_keys(value, parent_key=None):
    result = {}
    if not isinstance(value, dict):
        return result
    for k, v in value.items():
        new_key = f"{parent_key}.{k}" if parent_key else str(k)
        if isinstance(v, dict):
            result.update(_flatten_nested_keys(v, new_key))
        else:
            result[new_key] = v
    return result


def _normalize_subdoc(value):
    if not isinstance(value, dict):
        return value
    normalized = {}
    for k, v in value.items():
        if isinstance(v, dict):
            normalized.update(_flatten_nested_keys(v, str(k)))
        else:
            normalized[str(k)] = v
    return normalized


def _flatten_doc(doc: dict):
    flat = {}
    flat["_id"] = str(doc.get("_id")) if doc.get("_id") else None
    for container in ("identity", "salary", "attendance", "earnings", "deductions", "contributions"):
        value = doc.get(container)
        if isinstance(value, dict):
            flat.update(_normalize_subdoc(value))
    for k, v in doc.items():
        if k in {"_id", "identity", "salary", "attendance", "earnings", "deductions", "contributions"}:
            continue
        flat[k] = v

    # Recursively flatten any nested dictionaries that might have been created by MongoDB's dot-notation
    fully_flat = _flatten_nested_keys(flat)

    if "net_pay" in fully_flat and "Net Pay" not in fully_flat:
        fully_flat["Net Pay"] = fully_flat["net_pay"]
    if "Net Pay" in fully_flat and "net_pay" in fully_flat:
        del fully_flat["net_pay"]
    return fully_flat


def _row_filter(name: str, row_id: str):
    from bson.objectid import ObjectId
    # Check if row_id is a 24-char hex string (MongoDB _id)
    if len(row_id) == 24 and all(c in '0123456789abcdefABCDEF' for c in row_id):
        return {"_id": ObjectId(row_id)}
    
    if name == "payroll_records":
        if row_id.isdigit():
            return {"_row_id": int(row_id)}
        return {"emp_id": row_id}
    try:
        return {"_row_id": int(row_id)}
    except ValueError:
        return {"_row_id": row_id}


def _map_payroll_updates(name: str, to_write: dict, stored: dict = None):
    if name != "payroll_records":
        return to_write

    mapped = {}
    earnings_updates = {}
    deductions_updates = {}
    contributions_updates = {}
    salary_updates = {}
    attendance_updates = {}
    identity_updates = {}

    for key, value in to_write.items():
        if key.startswith("EARNING -"):
            earnings_updates[key] = value
            mapped[key] = value
        elif key.startswith("Deductions -"):
            deductions_updates[key] = value
            mapped[key] = value
        elif key == "Net Pay":
            mapped["net_pay"] = value
            mapped["Net Pay"] = value
        elif key in ("ATTENDANCE - Present Days", "ATTENDANCE - Pay Days"):
            attendance_updates[key] = value
            mapped[key] = value
        elif key.startswith("CONTRIBUTION -"):
            contributions_updates[key] = value
            mapped[key] = value
        elif key in ("emp_id", "emp_name", "location", "warehouse", "month", "year", "status"):
            mapped[key] = value
        elif key == "Mobile Number" or key == "mobile_number":
            mapped["mobile_number"] = value
            identity_updates["Mobile Number"] = value
            identity_updates["mobile_number"] = value
        elif key.startswith("FIXED -") or key in ("Fixed SC", "Uniform", "Shoes", "T Shirt", "Fixed Gross"):
            salary_updates[key] = value
            mapped[key] = value
        else:
            identity_updates[key] = value
            mapped[key] = value

    def _merge(subdoc_name, updates):
        if not updates:
            return
        existing = {}
        if stored and isinstance(stored.get(subdoc_name), dict):
            existing.update(stored[subdoc_name])
        existing.update(updates)
        mapped[subdoc_name] = existing

    _merge("earnings", earnings_updates)
    _merge("deductions", deductions_updates)
    _merge("contributions", contributions_updates)
    _merge("salary", salary_updates)
    _merge("attendance", attendance_updates)
    _merge("identity", identity_updates)
    return mapped


@app.get("/collections/{name}")
def get_collection(name: str):
    if name not in db.list_collection_names():
        raise HTTPException(404, "Unknown collection")
    sort_field = "_row_id" if name != "payroll_records" else "emp_id"
    docs = list(db[name].find().sort(sort_field, 1))
    if not docs:
        return {"columns": [], "rows": []}

    flat_docs = [_flatten_doc(d) for d in docs]
    columns, seen = [], set()
    for d in flat_docs:
        for k in d.keys():
            if k not in HIDDEN_FIELDS and k not in seen:
                seen.add(k)
                columns.append(k)

    rows = [{"_row_id": d.get("_row_id", d.get("emp_id")), **{c: d.get(c) for c in columns}} for d in flat_docs]
    return {"columns": columns, "rows": rows}


@app.get("/data/months")
def get_available_months(location: str = None, warehouse: str = None):
    """Return distinct (month, year) pairs from payroll records only."""
    query = {}
    if location:  query["location"]  = location
    if warehouse: query["warehouse"] = warehouse

    pairs = set()
    for doc in db["payroll_records"].find(query, {"month": 1, "year": 1}):
        if doc.get("month") and doc.get("year"):
            pairs.add((int(doc["month"]), int(doc["year"])))

    result = sorted(pairs, key=lambda x: (x[1], x[0]))
    return {"months": [{"month": m, "year": y} for m, y in result]}


@app.get("/events")
async def sse_events():
    return StreamingResponse(
        broadcaster.listen(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive"
        }
    )

@app.get("/payroll_record/{location}/{warehouse}/{month}/{year}/{emp_id}")
def get_payroll_record_endpoint(location: str, warehouse: str, month: int, year: int, emp_id: str):
    doc = db["payroll_records"].find_one({
        "location": location,
        "warehouse": warehouse,
        "month": month,
        "year": year,
        "emp_id": emp_id
    })
    if not doc:
        return {}
    flat = _flatten_doc(doc)
    flat["_id"] = str(doc.get("_id"))
    flat["_collection"] = "payroll_records"
    flat["_row_id"] = emp_id
    return flat

@app.get("/data")
def get_filtered_data(location: str = None, warehouse: str = None,
                      month: int = None, year: int = None):
    """Combined rows across every collection matching location/warehouse/month/year."""
    columns, seen = [], set()
    rows = []

    names = [n for n in db.list_collection_names() if not n.startswith("system.")]
    matching = []

    # Check for uploaded sheet collections matching location/warehouse (and month/year if specified)
    sheet_matching = []
    for name in names:
        if name in ("payroll_records", "employee_master") or name.startswith("_"):
            continue
        sample = db[name].find_one()
        if not sample:
            continue
        if location and sample.get("_location") != location:
            continue
        if warehouse and sample.get("_warehouse") != warehouse:
            continue
        if month and sample.get("_upload_month") and sample.get("_upload_month") != month:
            continue
        if year and sample.get("_upload_year") and sample.get("_upload_year") != year:
            continue
        sheet_matching.append(name)

    if sheet_matching:
        matching = sheet_matching
    elif month and year:
        matching = ["payroll_records"] if "payroll_records" in names else []
    else:
        for name in names:
            if name == "employee_master" or name.startswith("_"):
                continue
            sample = db[name].find_one()
            if not sample:
                continue
            if name == "payroll_records":
                if location and sample.get("location") != location:
                    continue
                if warehouse and sample.get("warehouse") != warehouse:
                    continue
            else:
                if location and sample.get("_location") != location:
                    continue
                if warehouse and sample.get("_warehouse") != warehouse:
                    continue
            matching.append(name)

    for name in matching:
        if name == "payroll_records":
            query = {}
            if month and year:
                query["month"] = month
                query["year"] = year
            if location:
                query["location"] = location
            if warehouse:
                query["warehouse"] = warehouse
        else:
            query = {}
            if month:
                query["_upload_month"] = month
            if year:
                query["_upload_year"] = year

        sort_field = "_row_id" if name != "payroll_records" else "emp_id"
        for d in db[name].find(query).sort(sort_field, 1):
            flat = _flatten_doc(d)
            for k in flat.keys():
                if k not in HIDDEN_FIELDS and k not in seen:
                    seen.add(k)
                    columns.append(k)
            rows.append({
                "_collection": name,
                "_id": str(d.get("_id")) if d.get("_id") else "",
                "_row_id": d.get("_row_id", d.get("emp_id")),
                "_sheet": d.get("_sheet"), "_warehouse": d.get("_warehouse"),
                "_upload_month": d.get("_upload_month"), "_upload_year": d.get("_upload_year"),
                **{c: flat.get(c) for c in columns},
            })
    for r in rows:
        for c in columns:
            r.setdefault(c, None)
    return {"collections": matching, "columns": columns, "rows": rows}


@app.post("/collections/{name}/rows")
def add_row(name: str, payload: dict = Body(...)):
    if name not in db.list_collection_names():
        raise HTTPException(404, "Unknown collection")
    data = {k: v for k, v in payload.items() if k not in HIDDEN_FIELDS}
    if not data:
        raise HTTPException(400, "No data provided")
    sample = db[name].find_one()
    last_id = db[name].find_one(sort=[("_row_id", -1)])
    next_id = (last_id["_row_id"] + 1) if last_id else 1
    now = datetime.utcnow()
    calculated = recalculate(data)
    doc = {**calculated, "_row_id": next_id,
           "_location": sample.get("_location") if sample else None,
           "_warehouse": sample.get("_warehouse") if sample else None,
           "_sheet": sample.get("_sheet") if sample else None,
           "_source_file": "manual",
           "_upload_month": now.month, "_upload_year": now.year}
    db[name].insert_one(doc)
    return {"ok": True, "_row_id": next_id}


@app.put("/collections/{name}/{row_id}")
def save_row(name: str, row_id: str, payload: dict = Body(...)):
    """Recalculate and save. Accepts edits to Fixed and Employee Info fields,
    rejects edits to calculated/attendance fields, runs engine, saves all."""
    if name not in db.list_collection_names():
        raise HTTPException(404, "Unknown collection")
    import re as _re
    from salary_calc import COLUMN_MAP, recalculate

    def _clean_key(k):
        return _re.sub(r'[\s]+', ' ', str(k)).strip()

    stored = db[name].find_one(_row_filter(name, row_id))
    if not stored:
        raise HTTPException(404, "row not found")

    flattened = _flatten_doc(stored)
    base = {_clean_key(k): v for k, v in flattened.items() if k != "_id"}

    # Frontend read-only columns (Attendance + Calculated)
    FRONTEND_READONLY = {
        'ATTENDANCE - Present Days', 'ATTENDANCE - Holi day', 'ATTENDANCE - Pay Days', 'ATTENDANCE - OT Hours',
        'EARNING - Basic', 'EARNING - DA', 'EARNING - Other Allows', 'EARNING - Leave With wages', 
        'EARNING - Bonus @8.33%', 'EARNING - HRA', 'EARNING - Total',
        'Net Pay',
        'Deductions - PF 12%', 'Deductions - ESIC 0.75%', 'Deductions - PT', 'Deductions - Adv', 'Deductions - Total Deduction',
        'CONTRIBUTION - EPF @ 13%', 'CONTRIBUTION - ESIC @ 3.25%', 'CONTRIBUTION - Total Employer Contribution',
        'CONTRIBUTION - CTC', 'CONTRIBUTION - Service Charge', 'CONTRIBUTION - Uniform Charges',
        'CONTRIBUTION - T Shirt', 'CONTRIBUTION - Shoes', 'CONTRIBUTION - Total CTC',
        'CONTRIBUTION - GST @18%', 'CONTRIBUTION - Total Billing'
    }

    # 1. Apply editable fields from payload to base
    editable_updates = {}
    manual_overrides = stored.get("_manual_overrides", []) if stored else []
    for k, v in payload.items():
        ck = _clean_key(k)
        if ck not in FRONTEND_READONLY and ck not in {_clean_key(h) for h in HIDDEN_FIELDS}:
            base[ck] = v
            editable_updates[ck] = v
            # If it's a known output column, record that the user manually edited it
            if ck in COLUMN_MAP.values() and ck not in manual_overrides:
                manual_overrides.append(ck)

    base["_manual_overrides"] = manual_overrides

    # 2. Run salary calculation on the updated base
    calculated = recalculate(base)
    calc_log = calculated.pop("__calc_log__", [])

    # 3. Write ALL changes (editable inputs + newly calculated outputs) back to Mongo
    to_write = {}
    for k, v in editable_updates.items():
        to_write[k] = v
    for k, v in calculated.items():
        if _clean_key(k) in {_clean_key(val) for val in COLUMN_MAP.values() if val}:
            to_write[_clean_key(k)] = v

    updates = {}
    if to_write:
        updates.update(_map_payroll_updates(name, to_write, stored=stored))
    
    updates["_manual_overrides"] = manual_overrides

    if not updates:
        raise HTTPException(400, "No columns to update")

    db[name].update_one(_row_filter(name, row_id), {"$set": updates})
    updated = db[name].find_one(_row_filter(name, row_id))
    clean = _flatten_doc(updated)
    clean.pop("_id", None)
    return {"ok": True, "row": clean, "calc_log": calc_log}


@app.post("/collections/{name}/{row_id}/recalculate")
def recalculate_row(name: str, row_id: str):
    if name not in db.list_collection_names():
        raise HTTPException(404, "Unknown collection")
    return save_row(name, row_id, {})


from payroll_records import ATTENDANCE_COLLECTION, PAYROLL_COLLECTION
from employee_master import MASTER_COLLECTION

@app.delete("/employee/{location}/{warehouse}/{emp_id}")
def delete_employee(location: str, warehouse: str, emp_id: str):
    att_res = db[ATTENDANCE_COLLECTION].delete_many({
        "location": location,
        "warehouse": warehouse,
        "emp_id": emp_id,
    })
    pay_res = db[PAYROLL_COLLECTION].delete_many({
        "location": location,
        "warehouse": warehouse,
        "emp_id": emp_id,
    })
    master_res = db[MASTER_COLLECTION].delete_many({
        "location": location,
        "warehouse": warehouse,
        "emp_id": emp_id,
    })
    return {
        "ok": True,
        "attendance_deleted": att_res.deleted_count,
        "payroll_deleted": pay_res.deleted_count,
        "master_deleted": master_res.deleted_count,
    }

@app.delete("/attendance/{location}/{warehouse}/{emp_id}")
def delete_employee_month(location: str, warehouse: str, emp_id: str, month: int, year: int):
    att_res = db[ATTENDANCE_COLLECTION].delete_many({
        "location": location,
        "warehouse": warehouse,
        "emp_id": emp_id,
        "month": month,
        "year": year
    })
    pay_res = db[PAYROLL_COLLECTION].delete_many({
        "location": location,
        "warehouse": warehouse,
        "emp_id": emp_id,
        "month": month,
        "year": year
    })
    return {
        "ok": True,
        "attendance_deleted": att_res.deleted_count,
        "payroll_deleted": pay_res.deleted_count,
    }


@app.patch("/collections/{name}/{row_id}")
def update_cell(name: str, row_id: str, payload: dict):
    column, value = payload.get("column"), payload.get("value")
    if column is None:
        raise HTTPException(400, "payload needs {column, value}")
    if column in HIDDEN_FIELDS:
        raise HTTPException(400, "that field isn't editable")
    result = db[name].update_one(_row_filter(name, row_id), {"$set": {column: value}})
    if result.matched_count == 0:
        raise HTTPException(404, "row not found")
    return {"ok": True}


# ---------------------------------------------------------------------------
# Excel Download endpoints
# ---------------------------------------------------------------------------
from fastapi.responses import StreamingResponse
import io
import openpyxl
from openpyxl.utils import get_column_letter

@app.get("/payroll/download")
def download_payroll(month: int, year: int, location: str = None, warehouse: str = None):
    query = {"month": month, "year": year}
    if location:  query["location"]  = location
    if warehouse: query["warehouse"] = warehouse

    docs = list(db["payroll_records"].find(query).sort("emp_id", 1))
    if not docs:
        raise HTTPException(404, "No payroll records found for the selected criteria.")

    flat_docs = [_flatten_doc(d) for d in docs]

    COLUMN_GROUPS = [
        ("Employee Details", [
            ("Emp ID", "emp_id"), ("Employee Name", "Employee Name"), ("Location", "location"), ("Warehouse", "warehouse"),
            ("Designation", "Designation"), ("Department", "Department"), ("Active Status", "Active Status"), ("DOJ", "DOJ"),
            ("ACCOUNT", "ACCOUNT"), ("Bank Name", "Bank Name"), ("IFSC", "IFSC"), ("UAN", "UAN"),
            ("ESIC", "ESIC"), ("PAN", "PAN"), ("Email", "Email"), ("Mobile Number", "Mobile Number")
        ]),
        ("Fixed", [
            ("Working Days", "FIXED - Working Days"), ("Basic", "FIXED - Basic"), ("DA", "FIXED - DA"), ("Other Allows", "FIXED - Other Allows"),
            ("Leave With wages", "FIXED - Leave With wages"), ("Bonus @8.33%", "FIXED - Bonus @8.33%"), ("HRA", "FIXED - HRA"), ("Total", "FIXED - Total")
        ]),
        ("Attendance", [
            ("Present Days", "ATTENDANCE - Present Days"), ("Holi day", "ATTENDANCE - Holi day"), ("Pay Days", "ATTENDANCE - Pay Days"), ("OT Hours", "ATTENDANCE - OT Hours")
        ]),
        ("Earned", [
            ("Basic", "EARNING - Basic"), ("DA", "EARNING - DA"), ("Other Allows", "EARNING - Other Allows"), ("Leave With wages", "EARNING - Leave With wages"),
            ("Bonus @8.33%", "EARNING - Bonus @8.33%"), ("HRA", "EARNING - HRA"), ("OT", "EARNING - OT Amount"), ("Total", "EARNING - Total")
        ]),
        ("Deductions", [
            ("PF 12%", "Deductions - PF 12%"), ("ESIC 0.75%", "Deductions - ESIC 0.75%"), ("PT", "Deductions - PT"), ("Adv", "Deductions - Adv"),
            ("Total Deduction", "Deductions - Total Deduction")
        ]),
        ("Employer Contribution", [
            ("EPF @ 13%", "CONTRIBUTION - EPF @ 13%"), ("ESIC @ 3.25%", "CONTRIBUTION - ESIC @ 3.25%"), ("Total Employer Contribution", "CONTRIBUTION - Total Employer Contribution")
        ]),
        ("CTC", [
            ("CTC", "CONTRIBUTION - CTC"), ("Service Charge", "CONTRIBUTION - Service Charge"), ("Uniform Charges", "CONTRIBUTION - Uniform Charges"), ("T Shirt", "CONTRIBUTION - T Shirt"),
            ("Shoes", "CONTRIBUTION - Shoes"), ("Total CTC", "CONTRIBUTION - Total CTC")
        ]),
        ("Net Pay", [
            ("Net Pay", "Net Pay")
        ])
    ]

    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Payroll_{month}_{year}"

    fill_group = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    fill_header = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    font_bold = Font(bold=True, color="FFFFFF")
    font_header = Font(bold=True, color="1E3A8A")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Row 1: Groups
    col_idx = 1
    for group_name, cols in COLUMN_GROUPS:
        start_col = col_idx
        end_col = col_idx + len(cols) - 1
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        cell = ws.cell(row=1, column=start_col, value=group_name)
        cell.fill = fill_group
        cell.font = font_bold
        cell.alignment = align_center
        cell.border = border_thin
        # Apply border to all merged cells
        for c in range(start_col, end_col + 1):
            ws.cell(row=1, column=c).border = border_thin
        col_idx = end_col + 1

    # Row 2: Columns
    col_idx = 1
    all_columns = []
    for group_name, cols in COLUMN_GROUPS:
        for header_name, db_key in cols:
            all_columns.append(db_key)
            cell = ws.cell(row=2, column=col_idx, value=header_name)
            cell.fill = fill_header
            cell.font = font_header
            cell.alignment = align_center
            cell.border = border_thin
            col_idx += 1

    # Data Rows
    row_idx = 3
    for d in flat_docs:
        for c_idx, c_name in enumerate(all_columns, start=1):
            val = d.get(c_name)
            if c_name == "emp_name" and not val:
                val = d.get("Employee Name", "")
            cell = ws.cell(row=row_idx, column=c_idx, value=val)
            cell.border = border_thin
            if isinstance(val, (int, float)):
                cell.number_format = '#,##0.00'
        row_idx += 1

    # Totals Row
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(COLUMN_GROUPS[0][1]))
    t_cell = ws.cell(row=row_idx, column=1, value="GRAND TOTAL")
    t_cell.font = Font(bold=True)
    t_cell.alignment = align_center
    t_cell.border = border_thin
    for c in range(1, len(COLUMN_GROUPS[0][1]) + 1):
        ws.cell(row=row_idx, column=c).border = border_thin

    for c_idx, c_name in enumerate(all_columns, start=1):
        if c_idx > len(COLUMN_GROUPS[0][1]): # Skip Employee Details
            col_letter = get_column_letter(c_idx)
            cell = ws.cell(row=row_idx, column=c_idx, value=f"=SUM({col_letter}3:{col_letter}{row_idx-1})")
            cell.font = Font(bold=True)
            cell.border = border_thin
            cell.number_format = '#,##0.00'

    # Auto-fit columns
    for col in ws.columns:
        max_len = 0
        for cell in col:
            if cell.row > 1 and cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 30)

    # Freeze panes
    ws.freeze_panes = "A3"

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    headers = {
        'Content-Disposition': f'attachment; filename="Payroll_{month}_{year}.xlsx"'
    }
    return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)

# ---------------------------------------------------------------------------
# Payroll Records endpoints
# ---------------------------------------------------------------------------

@app.post("/payroll/generate")
def generate_payroll(payload: dict = Body(...)):
    """
    Generate monthly payroll records from the employee master.
    Body: {month, year, location?, warehouse?}
    Existing records for that month are never overwritten.
    """
    month    = int(payload.get("month"))
    year     = int(payload.get("year"))
    location = payload.get("location", "")
    warehouse = payload.get("warehouse", "")

    if not month or not year:
        raise HTTPException(400, "month and year are required")

    result = generate_monthly_payroll(month, year, location, warehouse, db)
    return result


from bson.objectid import ObjectId

@app.post("/payroll/{payroll_id}/generate_payslip")
def generate_payslip_endpoint(payroll_id: str):
    record = db["payroll_records"].find_one({"_id": ObjectId(payroll_id)})
    if not record:
        raise HTTPException(404, "Payroll record not found")
    
    try:
        payslip_data = generate_and_upload_payslip(record, db)
        return {"success": True, "payslip": payslip_data}
    except Exception as e:
        print(f"Error generating payslip: {e}")
        raise HTTPException(500, f"Generation failed: {str(e)}")

@app.get("/payroll/{payroll_id}/download_payslip")
def download_payslip_endpoint(payroll_id: str):
    record = db["payroll_records"].find_one({"_id": ObjectId(payroll_id)})
    if not record:
        raise HTTPException(404, "Payroll record not found")
        
    payslip = record.get("payslip")
    if not payslip or not payslip.get("s3_key"):
        raise HTTPException(404, "Payslip PDF not generated yet")
        
    from s3_utils import generate_presigned_url
    url = generate_presigned_url(payslip["s3_key"])
    
    from fastapi.responses import RedirectResponse
    return RedirectResponse(url)

@app.post("/payroll/{payroll_id}/send_whatsapp")
def send_whatsapp_endpoint(payroll_id: str):
    record = db["payroll_records"].find_one({"_id": ObjectId(payroll_id)})
    if not record:
        raise HTTPException(404, "Payroll record not found")
        
    mobile = record.get("mobile_number")
    if not mobile:
        raise HTTPException(400, "Employee has no mobile number recorded")
        
    payslip = record.get("payslip")
    if not payslip or not payslip.get("s3_key"):
        raise HTTPException(400, "Payslip PDF not generated yet")
        
    # Re-generate the PDF bytes purely for the WhatsApp API payload
    try:
        from payslip_pdf import generate_payslip_pdf
        pdf_bytes = generate_payslip_pdf(record, db=db)
    except Exception as e:
        raise HTTPException(500, f"Failed to regenerate PDF: {str(e)}")
        
    success = send_payslip_whatsapp(
        phone_number=mobile,
        emp_name=record.get("emp_name", "Employee"),
        month=f"{record.get('month', '')}/{record.get('year', '')}",
        pdf_bytes=pdf_bytes,
        pdf_filename=payslip.get("file_name", "payslip.pdf")
    )
    
    if success:
        return {"success": True, "message": "WhatsApp sent successfully"}
    else:
        raise HTTPException(500, "WhatsApp delivery failed (check logs)")

@app.get("/download_bank_details")
def download_bank_details(month: int, year: int, location: str, warehouse: str):
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    import io
    from fastapi.responses import StreamingResponse
    from urllib.parse import quote

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bank Details"

    headers = ["Employee Name", "IFSC", "Account Number", "Location", "Warehouse"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="F97316", end_color="F97316", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    query = {"month": month, "year": year, "location": location, "warehouse": warehouse}
    docs = list(db["payroll_records"].find(query).sort("emp_id", 1))
    
    for row_num, d in enumerate(docs, 2):
        flat = _flatten_doc(d)
        emp_name = flat.get("Employee Name") or flat.get("emp_name") or ""
        ifsc = flat.get("IFSC") or ""
        acc = flat.get("ACCOUNT") or ""
        loc = flat.get("location") or ""
        wh = flat.get("warehouse") or ""
        
        ws.cell(row=row_num, column=1, value=emp_name)
        ws.cell(row=row_num, column=2, value=ifsc)
        ws.cell(row=row_num, column=3, value=str(acc) if acc else "")
        ws.cell(row=row_num, column=4, value=loc)
        ws.cell(row=row_num, column=5, value=wh)

    for i in range(1, 6):
        ws.column_dimensions[get_column_letter(i)].width = 25

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    filename = f"Bank_Details_{location}_{warehouse}_{month}_{year}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"}
    )

@app.get("/payroll")
def list_payroll(month: int, year: int,
                location: str = None, warehouse: str = None):
    """
    Return all payroll records for a given month/year.
    Optionally filter by location and warehouse.
    """
    records = get_payroll(month, year, location, warehouse, db)
    return {"records": records, "count": len(records)}


# ---------------------------------------------------------------------------
# Employee Master endpoints
# ---------------------------------------------------------------------------

@app.get("/employees")
def list_employees(location: str = None, warehouse: str = None, status: str = None):
    """List all employees from the permanent master."""
    query = {}
    if location:  query["location"]  = location
    if warehouse: query["warehouse"] = warehouse
    if status:    query["status"]    = status
    docs = list(db["employee_master"].find(query, {"_id": 0}).sort("emp_id", 1))
    return {"employees": docs, "count": len(docs)}


@app.post("/employees")
def create_employee(payload: dict = Body(...)):
    """
    Manually add a new employee to the master.
    Body: {
        emp_id, emp_name, location, warehouse,
        joined_month, joined_year,
        salary: { "FIXED - Basic": ..., ... },
        identity: { "Department": ..., ... }
    }
    Returns error if emp_id already exists for that location+warehouse.
    """
    required = ["emp_id", "emp_name", "location", "warehouse",
                "joined_month", "joined_year"]
    for f in required:
        if not payload.get(f):
            raise HTTPException(400, f"Missing required field: {f}")

    emp_id    = str(payload["emp_id"]).strip()
    location  = str(payload["location"]).strip()
    warehouse = str(payload["warehouse"]).strip()

    existing = db["employee_master"].find_one(
        {"emp_id": emp_id, "location": location, "warehouse": warehouse}
    )
    if existing:
        raise HTTPException(409, f"Employee {emp_id!r} already exists for "
                                 f"{location} / {warehouse}")

    doj = str(payload.get("doj") or (payload.get("identity", {}).get("DOJ") if isinstance(payload.get("identity"), dict) else "") or "").strip()
    doc = {
        "emp_id":       emp_id,
        "emp_name":     str(payload["emp_name"]).strip(),
        "location":     location,
        "warehouse":    warehouse,
        "status":       "active",
        "doj":          doj,
        "joined_month": int(payload["joined_month"]),
        "joined_year":  int(payload["joined_year"]),
        "salary":       payload.get("salary", {}),
        "identity":     payload.get("identity", {}),
        "last_updated": datetime.utcnow(),
        "last_source":  "manual",
    }
    db["employee_master"].insert_one(doc)
    
    # Automatically generate their payroll record for the month they joined,
    # so they immediately appear in the Salary Sheet without a manual "Generate" click.
    generate_monthly_payroll(
        month=doc["joined_month"],
        year=doc["joined_year"],
        location=location,
        warehouse=warehouse,
        db=db
    )
    
    return {"ok": True, "emp_id": emp_id}


@app.get("/employees/{emp_id}")
def get_employee(emp_id: str, location: str = None, warehouse: str = None):
    """Get a single employee from the master."""
    query = {"emp_id": emp_id}
    if location:  query["location"]  = location
    if warehouse: query["warehouse"] = warehouse
    doc = db["employee_master"].find_one(query, {"_id": 0})
    if not doc:
        raise HTTPException(404, f"Employee {emp_id!r} not found")
    return doc


# ---------------------------------------------------------------------------
# Attendance endpoints
# ---------------------------------------------------------------------------

ATTENDANCE_COLL = "_attendance"

PAY_STATUSES = {"p", "sl", "pl", "hp"}   # counts as present/paid
ABSENT_STATUS = "a"


def _att_key(emp_id: str, month: int, year: int, collection: str):
    return {"emp_id": emp_id, "month": month, "year": year, "collection": collection}


def _parse_attendance_date(key: str):
    if not isinstance(key, str):
        return None
    key = key.strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d", "%d %b %Y", "%d %B %Y"):
        try:
            return datetime.strptime(key, fmt).date()
        except ValueError:
            continue
    return None


from attendance_engine import get_payroll_dates, _calc_attendance


def _flatten_salary_doc(doc: dict):
    """Flatten payroll record fields for salary recalculation."""
    row = {}
    for k, v in doc.items():
        if k == "_id":
            continue
        if k in {"salary", "attendance"} and isinstance(v, dict):
            row.update(v)
        else:
            row[k] = v
    return row


@app.get("/attendance")
def get_attendance(collection: str, month: int, year: int):
    """Return all attendance records for a collection+month+year."""
    records = list(db[ATTENDANCE_COLL].find(
        {"month": month, "year": year, "collection": collection},
        {"_id": 0}
    ))
    return {"records": records}


def _ensure_payroll_record(emp_id: str, location: str, warehouse: str,
                           month: int, year: int):
    payroll_key = {
        "emp_id": {"$regex": f"^{emp_id}$", "$options": "i"},
        "location": {"$regex": f"^{location}$", "$options": "i"},
        "warehouse": {"$regex": f"^{warehouse}$", "$options": "i"},
        "month": month, 
        "year": year
    }
    payroll_doc = db["payroll_records"].find_one(payroll_key)
    if payroll_doc:
        return payroll_doc

    result = generate_monthly_payroll(month, year, location, warehouse, db)
    return db["payroll_records"].find_one(payroll_key)


@app.put("/attendance/{location}/{warehouse}/{emp_id}")
def save_attendance(location: str, warehouse: str, emp_id: str,
                    payload: dict = Body(...)):
    month = int(payload["month"])
    year  = int(payload["year"])
    new_days = {str(k): str(v).strip().upper() for k, v in payload.get("days", {}).items()}

    key = {"emp_id": emp_id, "location": location, "warehouse": warehouse}
    
    update_dict = {f"days.{k}": v for k, v in new_days.items()}
    update_dict["updated_at"] = datetime.utcnow()
    
    db[ATTENDANCE_COLL].update_one(key, {"$set": update_dict}, upsert=True)

    att_doc = db[ATTENDANCE_COLL].find_one(key)
    full_days = att_doc.get("days", {}) if att_doc else {}

    payroll_doc = _ensure_payroll_record(emp_id, location, warehouse, month, year)
    if payroll_doc:
        row = _flatten_salary_doc(payroll_doc)
        fixed_wd = row.get("FIXED - Working Days") or (row.get("identity", {}).get("FIXED - Working Days") if isinstance(row.get("identity"), dict) else None)
        try:
            fixed_wd = float(fixed_wd)
        except (ValueError, TypeError):
            fixed_wd = None
        master_emp = db["employee_master"].find_one({"emp_id": emp_id, "location": location, "warehouse": warehouse}) or {}
        doj_str = row.get("DOJ") or (row.get("identity", {}).get("DOJ") if isinstance(row.get("identity"), dict) else None) or master_emp.get("doj")
        leaving_date_str = row.get("leaving_date") or master_emp.get("leaving_date") or (master_emp.get("identity", {}).get("DOL") if isinstance(master_emp.get("identity"), dict) else None)
        present_days, absent_days, pay_days, lop, wd = _calc_attendance(emp_id, location, warehouse, month, year, full_days, fixed_wd, doj_str=doj_str, leaving_date_str=leaving_date_str)
        
        row["ATTENDANCE - Present Days"] = present_days
        row["ATTENDANCE - Pay Days"] = pay_days
        row["ATTENDANCE - LOP"] = lop
        
        calculated = recalculate(row)
        
        update_fields = {
            "attendance.ATTENDANCE - Present Days": present_days,
            "attendance.ATTENDANCE - Pay Days":     pay_days,
            "attendance.ATTENDANCE - LOP":          lop,
            "ATTENDANCE - Present Days":           present_days,
            "ATTENDANCE - Pay Days":               pay_days,
            "ATTENDANCE - LOP":                    lop,
            "updated_at": datetime.utcnow(),
        }
        for k, v in calculated.items():
            if k not in {"__calc_log__", "_id"} and not isinstance(v, dict):
                if k.startswith("EARNING -"):
                    update_fields[f"earnings.{k}"] = v
                    update_fields[k] = v
                elif k.startswith("Deductions -"):
                    update_fields[f"deductions.{k}"] = v
                    update_fields[k] = v
                elif k.startswith("CONTRIBUTION -"):
                    update_fields[f"contributions.{k}"] = v
                    update_fields[k] = v
                elif k == "Net Pay":
                    update_fields["net_pay"] = v
                    update_fields["Net Pay"] = v
        if update_fields:
            db["payroll_records"].update_one({"_id": payroll_doc["_id"]}, {"$set": update_fields})
            print(f"[SSE BROADCAST] Firing attendance save event for emp_id: {emp_id}")
            for q in broadcaster.queues:
                q.put_nowait(emp_id)

    return {"ok": True}


@app.delete("/employee/{location}/{warehouse}/{emp_id}")
def delete_employee(location: str, warehouse: str, emp_id: str, scope: str = "all", month: int = None, year: int = None):
    if scope == "all":
        db["employee_master"].delete_one({"emp_id": emp_id, "location": location, "warehouse": warehouse})
        db[ATTENDANCE_COLL].delete_many({"emp_id": emp_id, "location": location, "warehouse": warehouse})
        db["payroll_records"].delete_many({"emp_id": emp_id, "location": location, "warehouse": warehouse})
        
        # Also clean up any dynamic manual collections
        for coll in db.list_collection_names():
            if not coll.startswith("_") and coll not in {"employee_master", "payroll_records", "system.indexes"}:
                db[coll].delete_many({"emp_id": emp_id, "location": location, "warehouse": warehouse})
    elif scope == "month":
        if not month or not year:
            raise HTTPException(400, "Month and year required for scope=month")
        
        db["payroll_records"].delete_many({"emp_id": emp_id, "location": location, "warehouse": warehouse, "month": month, "year": year})
        
        period_dates = get_payroll_dates(month, year)
        unset_dict = {f"days.{d}": "" for d in period_dates}
        db[ATTENDANCE_COLL].update_one(
            {"emp_id": emp_id, "location": location, "warehouse": warehouse},
            {"$unset": unset_dict}
        )
    return {"ok": True, "deleted": emp_id, "scope": scope}


@app.get("/attendance/employees")
def get_attendance_employees(location: str, warehouse: str, month: int, year: int):
    from employee_master import is_employee_eligible_for_month
    query = {"location": location, "warehouse": warehouse}
    masters_raw = list(db["employee_master"].find(query, {"_id": 0}).sort("emp_id", 1))
    masters = [emp for emp in masters_raw if is_employee_eligible_for_month(emp, month, year)]

    # Fetch all attendance records for this location/warehouse in a single query
    att_cursor = db[ATTENDANCE_COLL].find(
        {"location": location, "warehouse": warehouse},
        {"_id": 0, "emp_id": 1, "days": 1}
    )
    att_map = {doc["emp_id"]: doc.get("days", {}) for doc in att_cursor}

    employees = []
    for emp in masters:
        emp_id = emp["emp_id"]
        full_days = att_map.get(emp_id, {})
        
        fixed_wd = emp.get("FIXED - Working Days")
        try:
            fixed_wd = float(fixed_wd)
        except (ValueError, TypeError):
            fixed_wd = None
            
        doj_str = emp.get("doj") or emp.get("DOJ") or (emp.get("identity", {}).get("DOJ") if isinstance(emp.get("identity"), dict) else None)
        leaving_date_str = emp.get("leaving_date") or (emp.get("identity", {}).get("DOL") or emp.get("identity", {}).get("leaving_date") if isinstance(emp.get("identity"), dict) else None)
        present_days, absent_days, pay_days, lop, wd = _calc_attendance(
            emp_id, location, warehouse, month, year, full_days, fixed_wd,
            doj_str=doj_str, leaving_date_str=leaving_date_str
        )
        
        employees.append({
            "emp_id":       emp_id,
            "emp_name":     emp.get("emp_name", ""),
            "department":   emp.get("Department", "") or (emp.get("identity", {}).get("Department", "") if isinstance(emp.get("identity"), dict) else ""),
            "designation":  emp.get("Designation", "") or (emp.get("identity", {}).get("Designation", "") if isinstance(emp.get("identity"), dict) else ""),
            "doj":          doj_str or "",
            "status":       emp.get("status", "active"),
            "leaving_date": leaving_date_str or "",
            "days":         full_days,
            "present_days": present_days,
            "absent_days":  absent_days,
            "pay_days":     pay_days,
            "lop":          lop,
            "working_days": wd
        })
    cfg = get_config()
    policy = cfg.get("attendance_policy", {})
    return {"employees": employees, "policy": policy}


@app.post("/employee/{location}/{warehouse}/{emp_id}/status")
def update_employee_status(location: str, warehouse: str, emp_id: str, payload: dict = Body(...)):
    status = str(payload.get("status", "")).strip().lower()
    if status not in ("active", "left", "discontinued"):
        raise HTTPException(400, "Invalid status. Must be 'active' or 'left' / 'discontinued'")

    month = int(payload.get("month", 0))
    year = int(payload.get("year", 0))
    leaving_date_str = str(payload.get("leaving_date", "")).strip()

    emp_filter = {"emp_id": emp_id, "location": location, "warehouse": warehouse}
    master = db["employee_master"].find_one(emp_filter)
    if not master:
        raise HTTPException(404, "Employee not found in master")

    now = datetime.utcnow()
    if status in ("left", "discontinued"):
        if not leaving_date_str:
            leaving_date_str = now.strftime("%Y-%m-%d")

        try:
            leaving_dt = datetime.strptime(leaving_date_str, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(400, "Invalid leaving_date format, must be YYYY-MM-DD")

        left_m = month or leaving_dt.month
        left_y = year or leaving_dt.year

        db["employee_master"].update_one(
            emp_filter,
            {"$set": {
                "status": "left",
                "leaving_date": leaving_date_str,
                "left_month": left_m,
                "left_year": left_y,
                "identity.DOL": leaving_date_str,
                "last_updated": now
            }}
        )

        # Update attendance in this cycle: mark dates after leaving_date as "L"
        period_dates = get_payroll_dates(left_m, left_y)
        att_updates = {}
        for d_str in period_dates:
            try:
                dt = datetime.strptime(d_str, "%Y-%m-%d").date()
                if dt > leaving_dt:
                    att_updates[f"days.{d_str}"] = "L"
            except ValueError:
                pass

        if att_updates:
            att_updates["updated_at"] = now
            db[ATTENDANCE_COLL].update_one(emp_filter, {"$set": att_updates}, upsert=True)

        # Recalculate attendance & payroll for this month
        att_doc = db[ATTENDANCE_COLL].find_one(emp_filter)
        full_days = att_doc.get("days", {}) if att_doc else {}
        payroll_doc = _ensure_payroll_record(emp_id, location, warehouse, left_m, left_y)
        if payroll_doc:
            row = _flatten_salary_doc(payroll_doc)
            fixed_wd = row.get("FIXED - Working Days") or (row.get("identity", {}).get("FIXED - Working Days") if isinstance(row.get("identity"), dict) else None)
            try:
                fixed_wd = float(fixed_wd)
            except (ValueError, TypeError):
                fixed_wd = None
            doj_str = master.get("doj") or (master.get("identity", {}).get("DOJ") if isinstance(master.get("identity"), dict) else None)
            present_days, absent_days, pay_days, lop, wd = _calc_attendance(
                emp_id, location, warehouse, left_m, left_y, full_days, fixed_wd,
                doj_str=doj_str, leaving_date_str=leaving_date_str
            )
            row["ATTENDANCE - Present Days"] = present_days
            row["ATTENDANCE - Pay Days"] = pay_days
            row["ATTENDANCE - LOP"] = lop
            calculated = recalculate(row)
            update_fields = {
                "attendance.ATTENDANCE - Present Days": present_days,
                "attendance.ATTENDANCE - Pay Days": pay_days,
                "attendance.ATTENDANCE - LOP": lop,
                "ATTENDANCE - Present Days": present_days,
                "ATTENDANCE - Pay Days": pay_days,
                "ATTENDANCE - LOP": lop,
                "status": "left",
                "leaving_date": leaving_date_str,
                "updated_at": now
            }
            for k, v in calculated.items():
                if k not in {"__calc_log__", "_id"} and not isinstance(v, dict):
                    if k.startswith("EARNING -"):
                        update_fields[f"earnings.{k}"] = v
                        update_fields[k] = v
                    elif k.startswith("Deductions -"):
                        update_fields[f"deductions.{k}"] = v
                        update_fields[k] = v
                    elif k.startswith("CONTRIBUTION -"):
                        update_fields[f"contributions.{k}"] = v
                        update_fields[k] = v
                    elif k == "Net Pay":
                        update_fields["net_pay"] = v
                        update_fields["Net Pay"] = v
            db["payroll_records"].update_one({"_id": payroll_doc["_id"]}, {"$set": update_fields})

        # Remove future month payroll records
        db["payroll_records"].delete_many({
            "emp_id": emp_id, "location": location, "warehouse": warehouse,
            "$or": [
                {"year": {"$gt": left_y}},
                {"year": left_y, "month": {"$gt": left_m}}
            ]
        })

    else:
        # Reactivate
        db["employee_master"].update_one(
            emp_filter,
            {
                "$set": {"status": "active", "last_updated": now},
                "$unset": {"leaving_date": "", "left_month": "", "left_year": "", "identity.DOL": ""}
            }
        )

    # Broadcast event
    for q in broadcaster.queues:
        q.put_nowait(emp_id)

    return {"ok": True, "emp_id": emp_id, "status": status, "leaving_date": leaving_date_str}


@app.patch("/employee/{location}/{warehouse}/{emp_id}/doj")
def update_employee_doj(location: str, warehouse: str, emp_id: str, payload: dict = Body(...)):
    doj_str = str(payload.get("doj", "")).strip()
    month = int(payload.get("month", 0))
    year = int(payload.get("year", 0))

    emp_filter = {"emp_id": emp_id, "location": location, "warehouse": warehouse}
    master = db["employee_master"].find_one(emp_filter)
    if not master:
        raise HTTPException(404, "Employee not found in master")

    now = datetime.utcnow()
    db["employee_master"].update_one(
        emp_filter,
        {"$set": {
            "doj": doj_str,
            "identity.DOJ": doj_str,
            "last_updated": now
        }}
    )

    if month and year:
        att_doc = db[ATTENDANCE_COLL].find_one(emp_filter)
        full_days = att_doc.get("days", {}) if att_doc else {}
        payroll_doc = _ensure_payroll_record(emp_id, location, warehouse, month, year)
        if payroll_doc:
            row = _flatten_salary_doc(payroll_doc)
            fixed_wd = row.get("FIXED - Working Days") or (row.get("identity", {}).get("FIXED - Working Days") if isinstance(row.get("identity"), dict) else None)
            try:
                fixed_wd = float(fixed_wd)
            except (ValueError, TypeError):
                fixed_wd = None
            leaving_date_str = master.get("leaving_date") or (master.get("identity", {}).get("DOL") if isinstance(master.get("identity"), dict) else None)
            present_days, absent_days, pay_days, lop, wd = _calc_attendance(
                emp_id, location, warehouse, month, year, full_days, fixed_wd,
                doj_str=doj_str, leaving_date_str=leaving_date_str
            )
            row["ATTENDANCE - Present Days"] = present_days
            row["ATTENDANCE - Pay Days"] = pay_days
            row["ATTENDANCE - LOP"] = lop
            calculated = recalculate(row)
            update_fields = {
                "attendance.ATTENDANCE - Present Days": present_days,
                "attendance.ATTENDANCE - Pay Days": pay_days,
                "attendance.ATTENDANCE - LOP": lop,
                "ATTENDANCE - Present Days": present_days,
                "ATTENDANCE - Pay Days": pay_days,
                "ATTENDANCE - LOP": lop,
                "DOJ": doj_str,
                "identity.DOJ": doj_str,
                "updated_at": now
            }
            for k, v in calculated.items():
                if k not in {"__calc_log__", "_id"} and not isinstance(v, dict):
                    if k.startswith("EARNING -"):
                        update_fields[f"earnings.{k}"] = v
                        update_fields[k] = v
                    elif k.startswith("Deductions -"):
                        update_fields[f"deductions.{k}"] = v
                        update_fields[k] = v
                    elif k.startswith("CONTRIBUTION -"):
                        update_fields[f"contributions.{k}"] = v
                        update_fields[k] = v
                    elif k == "Net Pay":
                        update_fields["net_pay"] = v
                        update_fields["Net Pay"] = v
            db["payroll_records"].update_one({"_id": payroll_doc["_id"]}, {"$set": update_fields})

    for q in broadcaster.queues:
        q.put_nowait(emp_id)

    return {"ok": True, "emp_id": emp_id, "doj": doj_str}



# ---------------------------------------------------------------------------

@app.get("/attendance/download")
def download_attendance(month: int, year: int, location: str, warehouse: str):
    from fastapi.responses import StreamingResponse
    import io
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import calendar
    from datetime import timedelta, datetime
    
    data = get_attendance_employees(location, warehouse, month, year)
    employees = data["employees"]
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Attendance_{month}_{year}"
    
    # Calculate dates: 26th of previous-previous month to 25th of previous month
    # Actually wait, in attendance.html:
    # const start = new Date(year, month - 2, 26);
    # const end = new Date(year, month - 1, 25);
    
    start_year = year
    start_month = month - 1
    if start_month < 1:
        start_month += 12
        start_year -= 1
        
    start_date = datetime(start_year, start_month, 26)
    
    end_year = year
    end_month = month
    if end_month < 1:
        end_month += 12
        end_year -= 1
        
    # Wait, the logic in frontend is: month - 2 to month - 1
    # Example: month=7 (July). start = Date(year, 5, 26) -> June 26th. 
    # end = Date(year, 6, 25) -> July 25th.
    
    from datetime import date
    
    # Let's just generate dates iteratively
    d_start_month = month - 1
    d_start_year = year
    if d_start_month < 1:
        d_start_month += 12
        d_start_year -= 1
        
    d_end_month = month
    d_end_year = year
    if d_end_month < 1:
        d_end_month += 12
        d_end_year -= 1
        
    start_d = date(d_start_year, d_start_month, 26)
    try:
        end_d = date(d_end_year, d_end_month, 25)
    except ValueError:
        end_d = date(d_end_year, d_end_month, 25)
        
    dates = []
    curr = start_d
    while curr <= end_d:
        dates.append(curr)
        curr += timedelta(days=1)
        
    # Headers
    headers = ["Emp ID", "Employee Name", "Department", "Designation", "DOJ"]
    date_strs = [d.strftime("%Y-%m-%d") for d in dates]
    headers.extend([str(d.day) for d in dates])
    headers.extend(["Working Days", "Present Days", "Paid Days", "LOP", "Absent Days"])
    
    ws.append(headers)
    
    # Header styles
    header_font = Font(bold=True)
    for col_idx, col_name in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        if col_idx <= 5:
            cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
        elif col_idx <= 5 + len(dates):
            # date columns
            is_abm = dates[col_idx-6].month == d_start_month
            if is_abm:
                cell.fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        else:
            cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
            
    for emp in employees:
        row = [
            emp["emp_id"],
            emp["emp_name"],
            emp["department"],
            emp["designation"],
            emp.get("doj", "")
        ]
        
        days_dict = emp.get("days", {})
        doj_str = emp.get("doj")
        doj_date = None
        if doj_str:
            for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d", "%d.%m.%Y", "%d %b %Y", "%d %B %Y"):
                try:
                    doj_date = datetime.strptime(str(doj_str).strip(), fmt).date()
                    break
                except ValueError:
                    pass

        leaving_str = emp.get("leaving_date")
        leaving_date = None
        if leaving_str:
            for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d", "%d.%m.%Y", "%d %b %Y", "%d %B %Y"):
                try:
                    leaving_date = datetime.strptime(str(leaving_str).strip(), fmt).date()
                    break
                except ValueError:
                    pass

        for d in dates:
            d_str = d.strftime("%Y-%m-%d")
            fallback = days_dict.get(str(d.day))
            val = days_dict.get(d_str, fallback) or ""
            if doj_date and d < doj_date:
                val = ""
            elif leaving_date and d > leaving_date:
                val = "L"
            elif val == "L":
                val = "L"
            elif not val:
                if d.weekday() == 6:
                    val = "WO"
                else:
                    val = "P"
            row.append(val)
            
        row.extend([
            emp.get("working_days", 0),
            emp.get("present_days", 0),
            emp.get("pay_days", 0),
            emp.get("lop", 0),
            emp.get("absent_days", 0)
        ])
        ws.append(row)
        
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=Attendance_{location}_{warehouse}_{month}_{year}.xlsx"}
    )


# Static files — keep LAST
# ---------------------------------------------------------------------------
# Resolve static files path relative to this file's directory
import os
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
STATIC_DIR = os.path.join(BASE_DIR, "static")
app.mount("/", StaticFiles(directory=STATIC_DIR, html=True), name="static")


from fastapi.responses import RedirectResponse
import zipfile
import io
import time

@app.get("/payroll/{row_id}/download_payslip")
def download_payslip_endpoint(row_id: str):
    print(f"HITTING DOWNLOAD ENDPOINT FOR {row_id}")
    try:
        obj_id = ObjectId(row_id)
    except Exception as e:
        print("ObjectId parse failed")
        raise HTTPException(400, "Invalid row_id format")
        
    record = db["payroll_records"].find_one({"_id": obj_id})
    if not record:
        print("Record not found in DB")
        raise HTTPException(404, "Payroll record not found")
        
    payslip = record.get("payslip", {})
    if not payslip.get("s3_key"):
        print("Payslip s3_key not found in record")
        raise HTTPException(404, "Payslip not generated yet")
    
    from s3_utils import generate_presigned_url
    signed_url = generate_presigned_url(record["payslip"]["s3_key"])
    return RedirectResponse(signed_url)

@app.post("/payroll/generate_all")
def generate_all_payslips(payload: dict = Body(...)):
    month = payload.get("month")
    year = payload.get("year")
    location = payload.get("location")
    warehouse = payload.get("warehouse")
    if not month or not year:
        raise HTTPException(400, "Month and year required")
    
    query = {"month": int(month), "year": int(year)}
    if location: query["location"] = location
    if warehouse: query["warehouse"] = warehouse
        
    records = list(db["payroll_records"].find(query))
    if not records:
        return {"status": "success", "generated": 0}
        
    from payslip_pdf import generate_and_upload_payslip
    generated = 0
    for record in records:
        try:
            generate_and_upload_payslip(record, db)
            generated += 1
        except Exception as e:
            print(f"Error generating for {record.get('emp_id')}: {e}")
            
    return {"status": "success", "generated": generated}

@app.get("/payroll/download_all")
def download_all_payslips(month: int, year: int, location: str = None, warehouse: str = None):
    query = {"month": month, "year": year}
    if location: query["location"] = location
    if warehouse: query["warehouse"] = warehouse
        
    records = list(db["payroll_records"].find(query))
    
    zip_buffer = io.BytesIO()
    from s3_utils import s3, S3_BUCKET
    
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for r in records:
            if r.get("payslip", {}).get("s3_key"):
                try:
                    obj = s3.get_object(Bucket=S3_BUCKET, Key=r["payslip"]["s3_key"])
                    pdf_bytes = obj["Body"].read()
                    filename = r["payslip"].get("file_name", f"{r.get('emp_id', 'payslip')}.pdf")
                    zip_file.writestr(filename, pdf_bytes)
                except Exception as e:
                    print(f"Failed to fetch {r['payslip']['s3_key']}: {e}")
                    
    zip_buffer.seek(0)
    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        zip_buffer,
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename=payslips_{month}_{year}.zip"}
    )

@app.post("/payroll/send_whatsapp_all")
def send_whatsapp_all(payload: dict = Body(...)):
    month = payload.get("month")
    year = payload.get("year")
    location = payload.get("location")
    warehouse = payload.get("warehouse")
    if not month or not year:
        raise HTTPException(400, "Month and year required")
        
    query = {"month": int(month), "year": int(year)}
    if location: query["location"] = location
    if warehouse: query["warehouse"] = warehouse
        
    records = list(db["payroll_records"].find(query))
    from app import send_whatsapp_endpoint
    sent = 0
    failed = 0
    for record in records:
        if str(record["_id"]):
            mob = record.get("mobile_number") or record.get("identity", {}).get("Mobile Number")
            if mob and record.get("payslip", {}).get("s3_key"):
                try:
                    res = send_whatsapp_endpoint(str(record["_id"]))
                    if res.get("status") == "success":
                        sent += 1
                    else:
                        failed += 1
                except:
                    failed += 1
                time.sleep(0.5)
                
    return {"status": "success", "sent": sent, "failed": failed}


# ---------------------------------------------------------------------------
# Salary Fitment / Increment Letter endpoints
# ---------------------------------------------------------------------------
from fitment_letter import (
    calculate_fitment_side,
    generate_fitment_excel,
    generate_fitment_pdf,
    send_fitment_letter_whatsapp,
)


@app.get("/fitment_letter/employee/{emp_id}")
def get_fitment_employee_data(emp_id: str, location: str = None, warehouse: str = None):
    """
    Fetch employee details and salary history to pre-populate Fitment Letter modal.
    """
    query = {"emp_id": emp_id}
    if location:
        query["location"] = location
    if warehouse:
        query["warehouse"] = warehouse

    master = db["employee_master"].find_one(query) or {}
    master_flat = _flatten_doc(master) if master else {}

    # Find the most recent payroll records for this employee
    records = list(
        db["payroll_records"]
        .find({"emp_id": emp_id})
        .sort([("year", -1), ("month", -1)])
        .limit(2)
    )

    current_rec = _flatten_doc(records[0]) if records else {}
    prev_rec = _flatten_doc(records[1]) if len(records) > 1 else {}

    # Extract info with fallbacks
    emp_name = current_rec.get("emp_name") or master_flat.get("emp_name") or current_rec.get("Employee Name") or ""
    designation = current_rec.get("Designation") or master_flat.get("Designation") or ""
    department = current_rec.get("Department") or current_rec.get("warehouse") or master_flat.get("warehouse") or master_flat.get("Department") or ""
    loc = current_rec.get("location") or master_flat.get("location") or "Bangalore"
    doj = current_rec.get("DOJ") or master_flat.get("DOJ") or ""
    mobile = current_rec.get("mobile_number") or current_rec.get("Mobile Number") or master_flat.get("mobile_number") or master_flat.get("Mobile Number") or ""

    # Current / New salary inputs
    new_basic = float(current_rec.get("FIXED - Basic") or master_flat.get("FIXED - Basic") or 0)
    new_da = float(current_rec.get("FIXED - DA") or master_flat.get("FIXED - DA") or 0)
    new_hra = float(current_rec.get("FIXED - HRA") or master_flat.get("FIXED - HRA") or 0)

    # Previous salary inputs (from earlier month or slightly lower if none)
    if prev_rec:
        prev_basic = float(prev_rec.get("FIXED - Basic") or new_basic)
        prev_da = float(prev_rec.get("FIXED - DA") or new_da)
        prev_hra = float(prev_rec.get("FIXED - HRA") or new_hra)
    else:
        prev_basic = new_basic
        prev_da = new_da
        prev_hra = new_hra

    now = datetime.utcnow()
    curr_yr = now.year % 100
    prev_yr = (now.year - 1) % 100
    prev_label = f"{prev_yr:02d}-{curr_yr:02d}"
    next_yr = (now.year + 1) % 100
    new_label = f"{curr_yr:02d}-{next_yr:02d}"

    return {
        "emp_id": emp_id,
        "emp_name": emp_name,
        "designation": designation,
        "department": department,
        "warehouse": department,
        "location": loc,
        "doj": doj,
        "mobile_number": mobile,
        "company_name": "RS MAN- TECH",
        "prev_year_label": prev_label,
        "new_year_label": new_label,
        "prev_basic": prev_basic,
        "prev_da": prev_da,
        "prev_hra": prev_hra,
        "new_basic": new_basic,
        "new_da": new_da,
        "new_hra": new_hra,
    }


@app.post("/fitment_letter/calculate")
def calculate_fitment_endpoint(payload: dict = Body(...)):
    loc = payload.get("location", "Bangalore")
    wh = payload.get("warehouse", payload.get("department", ""))
    
    prev_calc = calculate_fitment_side(
        basic=payload.get("prev_basic", 0),
        da=payload.get("prev_da", 0),
        hra=payload.get("prev_hra", 0),
        location=loc,
        warehouse=wh
    )
    new_calc = calculate_fitment_side(
        basic=payload.get("new_basic", 0),
        da=payload.get("new_da", 0),
        hra=payload.get("new_hra", 0),
        location=loc,
        warehouse=wh
    )
    return {"prev": prev_calc, "new": new_calc}


@app.post("/fitment_letter/download_excel")
def download_fitment_excel_endpoint(payload: dict = Body(...)):
    from urllib.parse import quote
    try:
        excel_bytes = generate_fitment_excel(payload)
        emp_name = payload.get("emp_name", "Employee").replace(" ", "_")
        filename = f"Fitment_Letter_{emp_name}_{payload.get('new_year_label', '26-27')}.xlsx"
        
        return StreamingResponse(
            io.BytesIO(excel_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"}
        )
    except Exception as e:
        raise HTTPException(500, f"Excel generation failed: {str(e)}")


@app.post("/fitment_letter/download_pdf")
def download_fitment_pdf_endpoint(payload: dict = Body(...)):
    from urllib.parse import quote
    try:
        pdf_bytes = generate_fitment_pdf(payload)
        emp_name = payload.get("emp_name", "Employee").replace(" ", "_")
        filename = f"Fitment_Letter_{emp_name}_{payload.get('new_year_label', '26-27')}.pdf"
        
        return StreamingResponse(
            io.BytesIO(pdf_bytes),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"}
        )
    except Exception as e:
        raise HTTPException(500, f"PDF generation failed: {str(e)}")


@app.post("/fitment_letter/send_whatsapp")
def send_fitment_whatsapp_endpoint(payload: dict = Body(...)):
    mobile = payload.get("mobile_number") or payload.get("mobile")
    if not mobile:
        raise HTTPException(400, "Employee mobile number is required")
        
    try:
        success = send_fitment_letter_whatsapp(str(mobile), payload)
        if success:
            return {"success": True, "message": f"Fitment Letter sent successfully to {mobile}"}
        else:
            raise HTTPException(500, "WhatsApp delivery failed (check logs/template)")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"WhatsApp dispatch failed: {str(e)}")

